import requests
from bs4 import BeautifulSoup
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook



def city(url):




    headers = {'Accept': '*/*',
               'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64;'
                             ' x64) AppleWebKit/537.36 '
                             '(KHTML, like Gecko) Chrome/104.0.5112.124 '
                             'YaBrowser/22.9.3.888 Yowser/2.5 Safari/537.36'
               }

    r = requests.get(url, headers=headers)
    soup = BeautifulSoup(r.text, 'lxml')

    streets = soup.find_all('div', class_="row")
    streets = streets[-1].find_all('a')
    for i in streets:

        req2 = requests.get(f"https://kladr-rf.ru{i.get('href')}", headers=headers)
        soup2 = BeautifulSoup(req2.text, 'lxml')
        numbers = soup2.find_all('table', class_="table table-bordered table-hover")

        if len(numbers) > 1:
            num = numbers[-1].find_all_next('td')
            count = 0
            listNum = []
            for j in range(len(num)):
                if count == 4:
                    count = 0
                if count == 0:
                    listNum.append(num[j].text)

                count += 1

            print(i.text, ','.join(listNum))


            fn = 'city-address.xlsx'

            wb = load_workbook(fn)
            ws = wb['Аркуш1']
            ws.append([i.text, ','.join(listNum)])
            wb.save(fn)
            wb.close()





def search(url):
    driver = webdriver.Chrome()
    driver.get(url)

    print("Введите населенный пункт: ")
    print()
    s = input()

    fn = 'city-address.xlsx'
    wb = load_workbook(fn)
    ws = wb['Аркуш1']
    ws.append([s])
    wb.save(fn)
    wb.close()



    user_name = driver.find_element(By.XPATH, '//*[@id="keyword"]')
    user_name.send_keys(s)
    button_log = driver.find_element(By.XPATH, '//*[@id="navbarsExample09"]/form/button')
    button_log.click()
    time.sleep(4)
    button_log1 = driver.find_element(By.XPATH, '//*[@id="L2AGLb"]/div')
    button_log1.click()
    button_log1 = driver.find_element(By.XPATH, '//*[@id="rso"]/div[1]/div/div/div[1]/div/a/h3')
    button_log1.click()

    time.sleep(2)
    with open("fls.html", "w", encoding="utf-8") as file:
        file.write(driver.page_source)
    driver.close()
    with open("fls.html", "r", encoding="utf-8") as file:
        file = file.read()
    soup = BeautifulSoup(file, "lxml")
    link = soup.find('ol', class_="breadcrumb").find_all('a')[-1].get('href')
    city(link)


def main():
    search('https://kladr-rf.ru')


if __name__ == '__main__':
    main()
